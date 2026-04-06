import os
import warnings
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def format_currency(amount):
    if amount == int(amount):
        return int(amount)
    return round(amount, 2)

def get_merge_start(ws, row, col):
    """Возвращает координаты верхней левой ячейки объединения"""
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return merged.min_row, merged.min_col
    return row, col

def safe_write(ws, row, col, value, as_text=False):
    """Безопасная запись"""
    if value is None:
        return
    target_row, target_col = get_merge_start(ws, row, col)
    cell = ws.cell(row=target_row, column=target_col)
    if as_text and isinstance(value, (int, float)):
        cell.value = str(int(value))
    else:
        cell.value = value

def write_digit(ws, row, col, digit):
    """Запись одной цифры"""
    if digit is None:
        return
    target_row, target_col = get_merge_start(ws, row, col)
    cell = ws.cell(row=target_row, column=target_col)
    cell.value = str(int(digit))

def write_letter(ws, row, col, letter):
    """Запись одной буквы"""
    if not letter:
        return
    target_row, target_col = get_merge_start(ws, row, col)
    cell = ws.cell(row=target_row, column=target_col)
    cell.value = letter

def write_oktmo_digits(ws, row, start_col, oktmo):
    """Запись ОКТМО (8 цифр) последовательно в ячейки"""
    oktmo_str = str(oktmo).strip()
    for i, digit in enumerate(oktmo_str):
        if i < 8 and digit.isdigit():
            write_digit(ws, row, start_col + i, int(digit))

def write_amount_digits(ws, row, start_col, amount):
    """Запись суммы последовательно в ячейки (до 12 цифр)"""
    amount_str = str(int(abs(amount)))
    for i, digit in enumerate(amount_str):
        if i < 12:
            write_digit(ws, row, start_col + i, int(digit))

def write_phone_by_letters(ws, phone):
    """Телефон: U27, W27, Y27, AA27, AC27, AE27, AG27, AI27, AK27, AM27, AO27"""
    phone_digits = ''.join(ch for ch in str(phone) if ch.isdigit())
    columns = [21, 23, 25, 27, 29, 31, 33, 35, 37, 39, 41]
    for i, digit in enumerate(phone_digits[:11]):
        if i < len(columns):
            write_digit(ws, 27, columns[i], int(digit))

def write_legal_name_by_letters(ws, name):
    """Название юрлица по буквам: A15, C15, E15..."""
    name_clean = ''.join(ch for ch in name.upper() if ch.isalpha() or ch == ' ')
    row = 15
    col = 1
    for char in name_clean:
        if char == ' ':
            char = ' '
        if col > 79:
            row = 17
            col = 1
        write_letter(ws, row, col, char)
        col += 2

def write_last_name_by_letters(ws, last_name):
    """Фамилия: B43, D43, F43..."""
    col = 2
    for char in last_name.upper():
        write_letter(ws, 43, col, char)
        col += 2

def write_first_name_by_letters(ws, first_name):
    """Имя: B45, D45, F45..."""
    col = 2
    for char in first_name.upper():
        write_letter(ws, 45, col, char)
        col += 2

def write_patronymic_by_letters(ws, patronymic):
    """Отчество: B47, D47, F47..."""
    col = 2
    for char in patronymic.upper():
        write_letter(ws, 47, col, char)
        col += 2


# ========== КУДиР ==========

def write_inn_digit_by_digit_kudir(ws, inn):
    inn_str = ''.join(ch for ch in str(inn) if ch.isdigit())
    positions = [1, 3, 5, 7, 9, 11, 13, 15, 17, 19, 21, 23]
    for i, digit in enumerate(inn_str):
        if i < len(positions):
            write_digit(ws, 28, positions[i], int(digit))

def fill_kudir_template(operations, template_path, output_path, inn, fio, ip_accounts, year=2025):
    wb = load_workbook(template_path)
    ws1 = wb["Лист1"]
    
    safe_write(ws1, 15, column_index_from_string('H'), year % 100, as_text=True)
    safe_write(ws1, 18, column_index_from_string('V'), fio)
    write_inn_digit_by_digit_kudir(ws1, inn)
    safe_write(ws1, 14, column_index_from_string('BB'), 1151085, as_text=True)
    
    today = datetime.now()
    safe_write(ws1, 15, column_index_from_string('BB'), today.year, as_text=True)
    safe_write(ws1, 15, column_index_from_string('BG'), today.month, as_text=True)
    safe_write(ws1, 15, column_index_from_string('BJ'), today.day, as_text=True)
    safe_write(ws1, 30, column_index_from_string('P'), "Доходы")
    
    row = 38
    for acc in ip_accounts:
        safe_write(ws1, row, 1, f"{acc['number']} {acc['bank']} БИК {acc['bik']}")
        row += 2
    
    wb.save(output_path)
    return sum(op['amount'] for op in operations)


# ========== ДЕКЛАРАЦИЯ ==========

def write_inn_digit_by_digit_titul(ws, inn):
    inn_str = ''.join(ch for ch in str(inn) if ch.isdigit())
    columns = [25, 27, 29, 31, 33, 35, 37, 39, 41, 43, 45, 47]
    for i, digit in enumerate(inn_str):
        if i < len(columns):
            write_digit(ws, 1, columns[i], int(digit))

def write_inn_digit_by_digit_section(ws, inn, start_col=13):
    """Универсальная запись ИНН с указанием начальной колонки"""
    inn_str = ''.join(ch for ch in str(inn) if ch.isdigit())
    columns = [start_col + i for i in range(12)]
    for i, digit in enumerate(inn_str):
        if i < len(columns):
            write_digit(ws, 1, columns[i], int(digit))

def write_tax_office_code(ws, inn):
    inn_str = ''.join(ch for ch in str(inn) if ch.isdigit())
    tax_code = inn_str[:4]
    columns = [27, 29, 31, 33]
    for i, digit in enumerate(tax_code):
        if i < len(columns):
            write_digit(ws, 13, columns[i], int(digit))

def write_place_of_registration_code(ws):
    write_digit(ws, 13, 75, 1)
    write_digit(ws, 13, 77, 2)
    write_digit(ws, 13, 79, 0)

def write_correction_number(ws):
    write_digit(ws, 11, 19, 0)

def write_tax_period_code(ws):
    write_digit(ws, 11, 53, 3)
    write_digit(ws, 11, 55, 4)

def write_report_year(ws, year):
    year_str = str(year)
    columns = [73, 75, 77, 79]
    for i, digit in enumerate(year_str):
        if i < len(columns):
            write_digit(ws, 11, columns[i], int(digit))

def write_director_last_name_titul(ws, last_name):
    target_row, target_col = get_merge_start(ws, 50, 8)
    cell = ws.cell(row=target_row, column=target_col)
    cell.value = last_name.upper()

def write_signature_date_titul(ws):
    today = datetime.now()
    day = str(today.day).zfill(2)
    month = str(today.month).zfill(2)
    year = str(today.year)
    
    write_digit(ws, 50, 22, int(day[0]))
    write_digit(ws, 50, 24, int(day[1]))
    write_digit(ws, 50, 28, int(month[0]))
    write_digit(ws, 50, 30, int(month[1]))
    write_digit(ws, 50, 34, int(year[0]))
    write_digit(ws, 50, 36, int(year[1]))
    write_digit(ws, 50, 38, int(year[2]))
    write_digit(ws, 50, 40, int(year[3]))

def write_director_last_name_section(ws, last_name, row=50, col=10):
    target_row, target_col = get_merge_start(ws, row, col)
    cell = ws.cell(row=target_row, column=target_col)
    cell.value = last_name.upper()

def write_signature_date_section(ws, row=50, col=22):
    today = datetime.now()
    date_str = today.strftime('%d.%m.%Y')
    target_row, target_col = get_merge_start(ws, row, col)
    cell = ws.cell(row=target_row, column=target_col)
    cell.value = date_str


def fill_declaration_template(operations, ens_data, template_path, output_excel, output_xml, inn, fio, oktmo, phone, tax_object):
    wb = load_workbook(template_path)
    
    # ========== ЛИСТ "Титул" (общий для обоих объектов) ==========
    if "Титул" not in wb.sheetnames:
        raise Exception(f"Лист 'Титул' не найден")
    
    ws_titul = wb["Титул"]
    
    write_inn_digit_by_digit_titul(ws_titul, inn)
    write_tax_office_code(ws_titul, inn)
    write_place_of_registration_code(ws_titul)
    write_correction_number(ws_titul)
    write_tax_period_code(ws_titul)
    write_report_year(ws_titul, 2025)
    write_legal_name_by_letters(ws_titul, f"ИНДИВИДУАЛЬНЫЙ ПРЕДПРИНИМАТЕЛЬ {fio}")
    
    if phone:
        write_phone_by_letters(ws_titul, phone)
    
    # Объект налогообложения на титуле
    write_digit(ws_titul, 29, 18, tax_object)
    
    fio_parts = fio.split()
    last_name = fio_parts[0] if len(fio_parts) > 0 else ""
    first_name = fio_parts[1] if len(fio_parts) > 1 else ""
    patronymic = fio_parts[2] if len(fio_parts) > 2 else ""
    
    if last_name:
        write_last_name_by_letters(ws_titul, last_name)
    if first_name:
        write_first_name_by_letters(ws_titul, first_name)
    if patronymic:
        write_patronymic_by_letters(ws_titul, patronymic)
    
    write_director_last_name_titul(ws_titul, last_name)
    write_signature_date_titul(ws_titul)
    
    # ========== ЛИСТ "Раздел 1.1" (для Доходов 6%) ==========
    if tax_object == 1:
        if "Раздел 1.1" not in wb.sheetnames:
            raise Exception(f"Лист 'Раздел 1.1' не найден")
        
        ws_s11 = wb["Раздел 1.1"]
        write_inn_digit_by_digit_section(ws_s11, inn, 13)
        write_oktmo_digits(ws_s11, 13, 26, oktmo)
        write_oktmo_digits(ws_s11, 18, 26, oktmo)
        write_oktmo_digits(ws_s11, 26, 26, oktmo)
        write_oktmo_digits(ws_s11, 34, 26, oktmo)
        write_director_last_name_section(ws_s11, last_name, 50, 10)
        write_signature_date_section(ws_s11, 50, 22)
    
    # ========== ЛИСТ "Раздел 1.2" (для Доходы минус расходы 15%) ==========
    elif tax_object == 2:
        if "Раздел 1.2" not in wb.sheetnames:
            raise Exception(f"Лист 'Раздел 1.2' не найден")
        
        ws_s12 = wb["Раздел 1.2"]
        write_inn_digit_by_digit_section(ws_s12, inn, 13)
        write_oktmo_digits(ws_s12, 13, 26, oktmo)
        write_oktmo_digits(ws_s12, 18, 26, oktmo)
        write_oktmo_digits(ws_s12, 26, 26, oktmo)
        write_oktmo_digits(ws_s12, 34, 26, oktmo)
        write_director_last_name_section(ws_s12, last_name, 50, 10)
        write_signature_date_section(ws_s12, 50, 22)
    
    # Расчет доходов
    quarterly = {1: 0.0, 2: 0.0, 3: 0.0, 4: 0.0}
    for op in operations:
        quarter = (op['date'].month - 1) // 3 + 1
        quarterly[quarter] += op['amount']
    
    total_income = sum(quarterly.values())
    
    if tax_object == 1:
        tax_rate = 6
    else:
        tax_rate = 15
    
    cum_income = {
        1: quarterly[1],
        2: quarterly[1] + quarterly[2],
        3: quarterly[1] + quarterly[2] + quarterly[3],
        4: total_income
    }
    
    cum_tax = {i: cum_income[i] * tax_rate / 100 for i in range(1, 5)}
    
    # Авансовые платежи из ЕНС
    usn_payments = ens_data.get('usn_payments', [])
    advance_payments = {1: 0.0, 2: 0.0, 3: 0.0}
    for payment in usn_payments:
        if payment['date']:
            month = payment['date'].month
            if month <= 3:
                advance_payments[1] += payment['amount']
            elif month <= 6:
                advance_payments[2] += payment['amount']
            elif month <= 9:
                advance_payments[3] += payment['amount']
    
    paid_in_2025 = any(d.year == 2025 for d in ens_data.get('insurance_paid_dates', []))
    insurance_paid = ens_data.get('insurance_paid', 0) if paid_in_2025 else 0
    cum_deductible = {i: min(cum_tax[i], insurance_paid) for i in range(1, 5)} if paid_in_2025 else {i: 0 for i in range(1, 5)}
    
    tax_payable = max(0, cum_tax[4] - cum_deductible[4] - advance_payments[1] - advance_payments[2] - advance_payments[3])
    
    # Заполнение разделов в зависимости от объекта
    if tax_object == 1:
        # Раздел 1.1
        if "Раздел 1.1" in wb.sheetnames:
            ws_s11 = wb["Раздел 1.1"]
            
            if advance_payments[1] > 0:
                write_amount_digits(ws_s11, 15, 26, advance_payments[1])
            else:
                write_digit(ws_s11, 15, 26, 0)
            
            if advance_payments[2] > 0:
                write_amount_digits(ws_s11, 20, 26, advance_payments[2])
            else:
                write_digit(ws_s11, 20, 26, 0)
            
            write_digit(ws_s11, 23, 26, 0)
            
            if advance_payments[3] > 0:
                write_amount_digits(ws_s11, 28, 26, advance_payments[3])
            else:
                write_digit(ws_s11, 28, 26, 0)
            
            write_digit(ws_s11, 31, 26, 0)
            
            if tax_payable > 0:
                write_amount_digits(ws_s11, 36, 26, tax_payable)
            else:
                write_digit(ws_s11, 36, 26, 0)
            
            if tax_payable < 0:
                write_amount_digits(ws_s11, 41, 26, abs(tax_payable))
            else:
                write_digit(ws_s11, 41, 26, 0)
        
        # Раздел 2.1.1
        if "Раздел 2.1.1" in wb.sheetnames:
            ws21 = wb["Раздел 2.1.1"]
            write_inn_digit_by_digit_section(ws21, inn, 13)
            write_digit(ws21, 11, 29, 2)
            
            write_amount_digits(ws21, 15, 29, cum_income[1])
            write_amount_digits(ws21, 17, 29, cum_income[2])
            write_amount_digits(ws21, 19, 29, cum_income[3])
            write_amount_digits(ws21, 21, 29, cum_income[4])
            
            write_amount_digits(ws21, 23, 29, tax_rate)
            write_amount_digits(ws21, 25, 29, tax_rate)
            write_amount_digits(ws21, 29, 29, tax_rate)
            
            write_amount_digits(ws21, 34, 29, cum_tax[1])
            write_amount_digits(ws21, 36, 29, cum_tax[2])
            write_amount_digits(ws21, 38, 29, cum_tax[3])
            write_amount_digits(ws21, 40, 29, cum_tax[4])
        
        # Раздел 2.1.1 (продолжение) - вычеты
        if "Раздел 2.1.1 (продолжение)" in wb.sheetnames:
            ws21_cont = wb["Раздел 2.1.1 (продолжение)"]
            write_amount_digits(ws21_cont, 11, 28, cum_deductible[1])
            write_amount_digits(ws21_cont, 14, 28, cum_deductible[2])
            write_amount_digits(ws21_cont, 17, 28, cum_deductible[3])
            write_amount_digits(ws21_cont, 20, 28, cum_deductible[4])
    
    elif tax_object == 2:
        # Для 15% пока заглушка (добавим позже)
        pass
    
    wb.save(output_excel)
    
    # XML
    xml = f'''<?xml version="1.0" encoding="UTF-8"?>
<Файл xmlns="urn:ФНС-СХД-Декл-УСН-2025-1">
    <Документ>
        <КНД>1152017</КНД>
        <ДатаДок>{datetime.now().strftime('%Y-%m-%d')}</ДатаДок>
        <НомКорр>0</НомКорр>
    </Документ>
    <НалогПериод>
        <НомерПериода>34</НомерПериода>
        <ОтчетныйГод>2025</ОтчетныйГод>
    </НалогПериод>
    <Налогоплательщик>
        <ИНН>{inn}</ИНН>
        <ИП>
            <ФИО>
                <Фамилия>{last_name}</Фамилия>
                <Имя>{first_name}</Имя>
                <Отчество>{patronymic}</Отчество>
            </ФИО>
        </ИП>
    </Налогоплательщик>
    <Показатели>
        <Раздел1_1>
            <ОКТМО>{oktmo}</ОКТМО>
            <СумАван010>{int(advance_payments[1])}</СумАван010>
            <СумАван020>{int(advance_payments[2])}</СумАван020>
            <СумАван040>{int(advance_payments[3])}</СумАван040>
            <СумАван070>0</СумАван070>
            <СумНал100>{int(tax_payable)}</СумНал100>
        </Раздел1_1>
        <Раздел2_1_1>
            <СумДоход110>{int(cum_income[1])}</СумДоход110>
            <СумДоход111>{int(cum_income[2])}</СумДоход111>
            <СумДоход112>{int(cum_income[3])}</СумДоход112>
            <СумДоход113>{int(cum_income[4])}</СумДоход113>
            <НалСтавка120>{tax_rate}</НалСтавка120>
            <СумИсчисНал130>{int(cum_tax[1])}</СумИсчисНал130>
            <СумИсчисНал131>{int(cum_tax[2])}</СумИсчисНал131>
            <СумИсчисНал132>{int(cum_tax[3])}</СумИсчисНал132>
            <СумИсчисНал133>{int(cum_tax[4])}</СумИсчисНал133>
            <СумУплНал140>{int(cum_deductible[1])}</СумУплНал140>
            <СумУплНал141>{int(cum_deductible[2])}</СумУплНал141>
            <СумУплНал142>{int(cum_deductible[3])}</СумУплНал142>
            <СумУплНал143>{int(cum_deductible[4])}</СумУплНал143>
        </Раздел2_1_1>
    </Показатели>
</Файл>'''
    
    with open(output_xml, 'w', encoding='utf-8') as f:
        f.write(xml)
    
    return tax_payable, total_income


def generate_report(operations, ens_data, output_dir, user_id, kudir_template, decl_template, inn, fio, oktmo, ip_accounts, phone, tax_object):
    kudir_path = os.path.join(output_dir, f"kudir_{user_id}.xlsx")
    total_income = fill_kudir_template(operations, kudir_template, kudir_path, inn, fio, ip_accounts)
    
    decl_excel = os.path.join(output_dir, f"declaration_{user_id}.xlsx")
    decl_xml = os.path.join(output_dir, f"declaration_{user_id}.xml")
    tax_payable, total_income = fill_declaration_template(
        operations, ens_data, decl_template, decl_excel, decl_xml, inn, fio, oktmo, phone, tax_object
    )
    
    return kudir_path, decl_excel, decl_xml, total_income, tax_payable